from PyQt6 import QtCore, QtGui, QtWidgets
import sys, os
import openpyxl
import operator

statistic = []
num = []
banlist = []

class Ui_Window(object):

    def setupUi(self, Window):
        Window.setObjectName("Window")
        Window.setEnabled(True)
        Window.resize(833, 520)
        Window.setMouseTracking(False)
        Window.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)
        Window.setWindowTitle("")
        Window.setStyleSheet("background-color: #075346;\n"
"font: 28 pt \"Bahnschrift Condensed\";\n"
"color: white")
        Window.setWindowTitle("Analytics")
        self.verticalLayoutWidget = QtWidgets.QWidget(parent=Window)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(210, 40, 381, 80))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.Header_Layout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.Header_Layout.setContentsMargins(0, 0, 0, 0)
        self.Header_Layout.setObjectName("Header_Layout")
        self.Header = QtWidgets.QLabel(parent=self.verticalLayoutWidget)
        self.Header.setStyleSheet("font: 28pt;")
        self.Header.setObjectName("Header")
        self.Header_Layout.addWidget(self.Header)
        self.stackedWidget = QtWidgets.QStackedWidget(parent=Window)
        self.stackedWidget.setGeometry(QtCore.QRect(20, 160, 800, 361))
        self.stackedWidget.setObjectName("stackedWidget")
        self.Page1 = QtWidgets.QWidget()
        self.Page1.setStyleSheet("font: 24pt;\n"
"")
        self.Page1.setObjectName("Page1")
        self.verticalLayoutWidget_2 = QtWidgets.QWidget(parent=self.Page1)
        self.verticalLayoutWidget_2.setGeometry(QtCore.QRect(10, 10, 289, 161))
        self.verticalLayoutWidget_2.setObjectName("verticalLayoutWidget_2")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_2)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.Label_sports = QtWidgets.QLabel(parent=self.verticalLayoutWidget_2)
        self.Label_sports.setStyleSheet("")
        self.Label_sports.setObjectName("Label_sports")
        self.verticalLayout.addWidget(self.Label_sports)
        self.Label_choose = QtWidgets.QLabel(parent=self.verticalLayoutWidget_2)
        self.Label_choose.setObjectName("Label_choose")
        self.verticalLayout.addWidget(self.Label_choose)
        self.verticalLayoutWidget_3 = QtWidgets.QWidget(parent=self.Page1)
        self.verticalLayoutWidget_3.setGeometry(QtCore.QRect(240, 10, 511, 71))
        self.verticalLayoutWidget_3.setObjectName("verticalLayoutWidget_3")
        self.Choose_sports = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_3)
        self.Choose_sports.setContentsMargins(0, 0, 0, 0)
        self.Choose_sports.setObjectName("Choose_sports")
        self.Sports_label1 = QtWidgets.QComboBox(parent=self.verticalLayoutWidget_3)
        self.Sports_label1.setStyleSheet("background: #B0B0B0;")
        self.Sports_label1.setObjectName("Sports_label1")
        self.Sports_label1.addItem("")
        self.Sports_label1.addItem("")
        self.Sports_label1.addItem("")
        self.Sports_label1.addItem("")
        self.Choose_sports.addWidget(self.Sports_label1)
        self.verticalLayoutWidget_4 = QtWidgets.QWidget(parent=self.Page1)
        self.verticalLayoutWidget_4.setGeometry(QtCore.QRect(240, 90, 511, 80))
        self.verticalLayoutWidget_4.setObjectName("verticalLayoutWidget_4")
        self.Choose_team = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_4)
        self.Choose_team.setContentsMargins(0, 0, 0, 0)
        self.Choose_team.setObjectName("Choose_team")
        self.Team = QtWidgets.QComboBox(parent=self.verticalLayoutWidget_4)
        self.Team.setStyleSheet("background: #B0B0B0;")
        self.Team.setObjectName("Team")
        self.Team.addItem("")
        self.Team.addItem("")
        self.Choose_team.addWidget(self.Team)
        self.Push_done = QtWidgets.QPushButton(parent=self.Page1)
        self.Push_done.setGeometry(QtCore.QRect(10, 210, 231, 71))
        self.Push_done.setStyleSheet("background: #0D9B83")
        self.Push_done.setObjectName("Push_done")
        self.footer = QtWidgets.QLabel(parent=self.Page1)
        self.footer.setGeometry(QtCore.QRect(560, 210, 211, 71))
        self.footer.setObjectName("footer")
        self.Error = QtWidgets.QLabel(parent=self.Page1)
        self.Error.setGeometry(QtCore.QRect(276, 212, 251, 71))
        self.Error.setObjectName("Error")
        self.stackedWidget.addWidget(self.Page1)
        self.Page2 = QtWidgets.QWidget()
        self.Page2.setStyleSheet("font: 24pt")
        self.Page2.setObjectName("Page2")
        self.horizontalLayoutWidget_2 = QtWidgets.QWidget(parent=self.Page2)
        self.horizontalLayoutWidget_2.setGeometry(QtCore.QRect(30, 10, 714, 80))
        self.horizontalLayoutWidget_2.setObjectName("horizontalLayoutWidget_2")
        self.Navigation_Layout = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget_2)
        self.Navigation_Layout.setContentsMargins(0, 0, 0, 0)
        self.Navigation_Layout.setObjectName("Navigation_Layout")
        self.Result = QtWidgets.QLabel(parent=self.horizontalLayoutWidget_2)
        self.Result.setObjectName("Result")
        self.Navigation_Layout.addWidget(self.Result)
        self.Full_time = QtWidgets.QLabel(parent=self.horizontalLayoutWidget_2)
        self.Full_time.setObjectName("Full_time")
        self.Navigation_Layout.addWidget(self.Full_time)
        self.horizontalLayoutWidget_3 = QtWidgets.QWidget(parent=self.Page2)
        self.horizontalLayoutWidget_3.setGeometry(QtCore.QRect(210, 260, 561, 91))
        self.horizontalLayoutWidget_3.setObjectName("horizontalLayoutWidget_3")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget_3)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.Setting_fullname = QtWidgets.QPushButton(parent=self.horizontalLayoutWidget_3)
        self.Setting_fullname.setStyleSheet("background: #0D9B83")
        self.Setting_fullname.setObjectName("Setting_fullname")
        self.horizontalLayout.addWidget(self.Setting_fullname)
        self.Come_back_button1 = QtWidgets.QPushButton(parent=self.horizontalLayoutWidget_3)
        self.Come_back_button1.setStyleSheet("background: #0D9B83")
        self.Come_back_button1.setObjectName("Come_back_button1")
        self.horizontalLayout.addWidget(self.Come_back_button1)
        self.verticalLayoutWidget_7 = QtWidgets.QWidget(parent=self.Page2)
        self.verticalLayoutWidget_7.setGeometry(QtCore.QRect(390, 100, 211, 151))
        self.verticalLayoutWidget_7.setObjectName("verticalLayoutWidget_7")
        self.Answers = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_7)
        self.Answers.setContentsMargins(0, 0, 0, 0)
        self.Answers.setObjectName("Answers")
        self.Result1 = QtWidgets.QLabel(parent=self.verticalLayoutWidget_7)
        self.Result1.setEnabled(True)
        self.Result1.setStyleSheet("background: #B0B0B0;\n"
"border-radius: 10%;")
        self.Result1.setText("")
        self.Result1.setObjectName("Result1")
        self.Answers.addWidget(self.Result1)
        self.Result2 = QtWidgets.QLabel(parent=self.verticalLayoutWidget_7)
        self.Result2.setStyleSheet("background: #B0B0B0;\n"
"border-radius: 10%;")
        self.Result2.setText("")
        self.Result2.setObjectName("Result2")
        self.Answers.addWidget(self.Result2)
        self.Result3 = QtWidgets.QLabel(parent=self.verticalLayoutWidget_7)
        self.Result3.setStyleSheet("background: #B0B0B0;\n"
"border-radius: 10%;")
        self.Result3.setText("")
        self.Result3.setObjectName("Result3")
        self.Answers.addWidget(self.Result3)
        self.Button_settings = QtWidgets.QPushButton(parent=self.Page2)
        self.Button_settings.setGeometry(QtCore.QRect(80, 260, 95, 95))
        self.Button_settings.setStyleSheet("border-image : url(Image/setting.png);")
        self.Button_settings.setObjectName("Button_settings")
        self.verticalLayoutWidget_18 = QtWidgets.QWidget(parent=self.Page2)
        self.verticalLayoutWidget_18.setGeometry(QtCore.QRect(30, 100, 291, 151))
        self.verticalLayoutWidget_18.setObjectName("verticalLayoutWidget_18")
        self.Answers_3 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_18)
        self.Answers_3.setContentsMargins(0, 0, 0, 0)
        self.Answers_3.setObjectName("Answers_3")
        self.Result1_3 = QtWidgets.QLabel(parent=self.verticalLayoutWidget_18)
        self.Result1_3.setEnabled(True)
        self.Result1_3.setStyleSheet("background: #B0B0B0;\n"
"border-radius: 10%;")
        self.Result1_3.setText("")
        self.Result1_3.setObjectName("Result1_3")
        self.Answers_3.addWidget(self.Result1_3)
        self.Result2_3 = QtWidgets.QLabel(parent=self.verticalLayoutWidget_18)
        self.Result2_3.setStyleSheet("background: #B0B0B0;\n"
"border-radius: 10%;")
        self.Result2_3.setText("")
        self.Result2_3.setObjectName("Result2_3")
        self.Answers_3.addWidget(self.Result2_3)
        self.Result3_3 = QtWidgets.QLabel(parent=self.verticalLayoutWidget_18)
        self.Result3_3.setStyleSheet("background: #B0B0B0;\n"
"border-radius: 10%;")
        self.Result3_3.setText("")
        self.Result3_3.setObjectName("Result3_3")
        self.Answers_3.addWidget(self.Result3_3)
        self.stackedWidget.addWidget(self.Page2)
        self.Page3 = QtWidgets.QWidget()
        self.Page3.setStyleSheet("font: 24pt;")
        self.Page3.setObjectName("Page3")
        self.verticalLayoutWidget_9 = QtWidgets.QWidget(parent=self.Page3)
        self.verticalLayoutWidget_9.setGeometry(QtCore.QRect(10, 40, 313, 231))
        self.verticalLayoutWidget_9.setObjectName("verticalLayoutWidget_9")
        self.Information = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_9)
        self.Information.setContentsMargins(0, 0, 0, 0)
        self.Information.setObjectName("Information")
        self.Fullname = QtWidgets.QLabel(parent=self.verticalLayoutWidget_9)
        self.Fullname.setObjectName("Fullname")
        self.Information.addWidget(self.Fullname)
        self.Add_a_student = QtWidgets.QLabel(parent=self.verticalLayoutWidget_9)
        self.Add_a_student.setObjectName("Add_a_student")
        self.Information.addWidget(self.Add_a_student)
        self.Sports_label2 = QtWidgets.QLabel(parent=self.verticalLayoutWidget_9)
        self.Sports_label2.setObjectName("Sports_label2")
        self.Information.addWidget(self.Sports_label2)
        self.verticalLayoutWidget_8 = QtWidgets.QWidget(parent=self.Page3)
        self.verticalLayoutWidget_8.setGeometry(QtCore.QRect(340, 30, 401, 251))
        self.verticalLayoutWidget_8.setObjectName("verticalLayoutWidget_8")
        self.Addition = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_8)
        self.Addition.setContentsMargins(0, 0, 0, 0)
        self.Addition.setObjectName("Addition")
        self.lineEdit = QtWidgets.QLineEdit(parent=self.verticalLayoutWidget_8)
        self.lineEdit.setStyleSheet("background: #B0B0B0;\n"
"border-radius: 10%;")
        self.lineEdit.setText("")
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.lineEdit.setClearButtonEnabled(True)
        self.Addition.addWidget(self.lineEdit)
        self.add_fullname = QtWidgets.QLineEdit(parent=self.verticalLayoutWidget_8)
        self.add_fullname.setAutoFillBackground(False)
        self.add_fullname.setStyleSheet("background: #B0B0B0;\n"
"border-radius: 10%;")
        self.add_fullname.setText("")
        self.add_fullname.setFrame(False)
        self.add_fullname.setDragEnabled(False)
        self.add_fullname.setClearButtonEnabled(True)
        self.add_fullname.setObjectName("add_fullname")
        self.add_fullname.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.Addition.addWidget(self.add_fullname)
        self.Choose_sports_label = QtWidgets.QComboBox(parent=self.verticalLayoutWidget_8)
        self.Choose_sports_label.setStyleSheet("background: #B0B0B0;\n"
"border-radius: 10%;")
        self.Choose_sports_label.setObjectName("Choose_sports_label")
        self.Choose_sports_label.addItem("")
        self.Choose_sports_label.addItem("")
        self.Addition.addWidget(self.Choose_sports_label)
        self.verticalLayoutWidget_10 = QtWidgets.QWidget(parent=self.Page3)
        self.verticalLayoutWidget_10.setGeometry(QtCore.QRect(10, 270, 311, 101))
        self.verticalLayoutWidget_10.setObjectName("verticalLayoutWidget_10")
        self.Is_addition_student = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_10)
        self.Is_addition_student.setContentsMargins(0, 0, 0, 0)
        self.Is_addition_student.setObjectName("Is_addition_student")
        self.pushButton_3 = QtWidgets.QPushButton(parent=self.verticalLayoutWidget_10)
        self.pushButton_3.setStyleSheet("background: #0D9B83")
        self.pushButton_3.setObjectName("pushButton_3")
        self.Is_addition_student.addWidget(self.pushButton_3)
        self.horizontalLayoutWidget_4 = QtWidgets.QWidget(parent=self.Page3)
        self.horizontalLayoutWidget_4.setGeometry(QtCore.QRect(340, 290, 401, 61))
        self.horizontalLayoutWidget_4.setObjectName("horizontalLayoutWidget_4")
        self.Come_back_page3 = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget_4)
        self.Come_back_page3.setContentsMargins(0, 0, 0, 0)
        self.Come_back_page3.setObjectName("Come_back_page3")
        self.Come_back_button2 = QtWidgets.QPushButton(parent=self.horizontalLayoutWidget_4)
        self.Come_back_button2.setStyleSheet("background: #0D9B83")
        self.Come_back_button2.setObjectName("Come_back_button2")
        self.Come_back_page3.addWidget(self.Come_back_button2)
        self.stackedWidget.addWidget(self.Page3)
        self.Logo = QtWidgets.QLabel(parent=Window)
        self.Logo.setGeometry(QtCore.QRect(585, 0, 170, 180))
        self.Logo.setMaximumSize(QtCore.QSize(175, 175))
        self.Logo.setStyleSheet("")
        self.Logo.setText("")
        self.Logo.setPixmap(QtGui.QPixmap("Image/Logo.png"))
        self.Logo.setScaledContents(True)
        self.Logo.setObjectName("Logo")

        self.retranslateUi(Window)
        self.stackedWidget.setCurrentIndex(1)
        QtCore.QMetaObject.connectSlotsByName(Window)

        # Устонавливаем перввую страницу.
        self.stackedWidget.setCurrentIndex(0)

        # Переход на предыдущую страницу
        self.Come_back_button1.clicked.connect(self.EventMouseComeBack)
        self.Come_back_button2.clicked.connect(self.EventMouseComeBack)
        # Переход на следующую страницу
        self.Push_done.clicked.connect(self.EventMousePush)
        self.Button_settings.clicked.connect(self.EventMousePush)


        self.pushButton_3.clicked.connect(self.Add_data_base)

        #Поиск лучшей даты
        self.Setting_fullname.clicked.connect(self.find_better_date)

    def find_better_date(self):
        ls = self.algor(imp="М2")
        print(ls)
        self.Result1_3.setText(str(ls[0][0]))
        self.Result2_3.setText(str(ls[1][0]))
        self.Result3_3.setText(str(ls[2][0]))
        self.Result1.setText(str(ls[0][1]))
        self.Result2.setText(str(ls[1][1]))
        self.Result3.setText(str(ls[2][1]))

    def Add_data_base(self):
        team = str(self.lineEdit.text())
        date = str(self.add_fullname.text())
        result = str(self.Choose_sports_label.currentText())
        section = str(self.Sports_label1.currentText())

        if not os.path.isfile(f'source/{section}.xlsx'):
                with open(f"source/{section}.xlsx", "w") as file:
                        print(file)

        if len(team) == 0 or len(date) == 0 or section == 'Выберите секцию':
                return None
        else:
                workbook = openpyxl.load_workbook(f"source/{section}.xlsx")
                sheet = workbook.active
                sheet["A1"] = "Команда"
                sheet["B1"] = "Дата"
                sheet["C1"] = "Результат"
                sheet.append([team, date, result])
                workbook.save(f"source/{section}.xlsx")

    def EventMouseComeBack(self):
            self.stackedWidget.setCurrentIndex(self.stackedWidget.currentIndex() - 1)

    def EventMousePush(self):
            self.stackedWidget.setCurrentIndex(self.stackedWidget.currentIndex() + 1)

    # def is_check_to_choose_page1(self):
    #         currentText1 = self.Sports_label1.currentText()
    #         currentText2 = self.Team.currentText()
    #         print(f"currentText1: {currentText1}, currentText2: {currentText2}")
    #
    #         if currentText1 == 'Выберите вид спорта' and currentText2 == 'Команда':
    #                 self.Error.setText("Выберите секцию и команду")
    #                 self.Error.setStyleSheet("font: 18pt; color: red;")
    #         else:
    #                 print("Условие не выполнено")


    def sortser(self, company):
            comp = []
            section = str(self.Sports_label1.currentText())
            workbook = openpyxl.load_workbook(f"Data (3).xlsx")
            sheet = workbook.active

            for i in range(2, sheet.max_row + 1):
                    date = sheet.cell(row=i, column=1)
                    date = date.value
                    if date == company:
                        comp.append(i)
            return comp

    # Функция приводящая получаемые из таблицы данные в удобные для обработки значения
    def statisticser(self,sorter):
            section = str(self.Sports_label1.currentText())
            workbook = openpyxl.load_workbook(f"Data (3).xlsx")
            sheet = workbook.active
            i = 0
            while i < len(sorter):
                    date = sheet.cell(row=sorter[i], column=2)
                    date = str(date.value)
                    vic = sheet.cell(row=sorter[i], column=3)
                    vic = str(vic.value)

                    # print(date)
                    date = date.split(' ')
                    date = date[0]
                    date = date.split('-')
                    date = str(date[1] + "." + date[2])
                    # Дата в Американском фармате(позволяет использовать встроенный sort)
                    statistic.append(date)
                    statistic.append(vic)
                    i = i + 1

            return statistic

    # Функция Получения вероятности победы

    def equalser(self, list, num):
            attemp = []
            word1 = list
            i = 0

            leng = len(list)
            while i < (leng - 1):
                    j = i
                    win = 0
                    while (j + 1 < leng):

                            if word1[j] == "Победа" or word1[j] == "Поражение":
                                    break
                            if word1[j] == word1[j + 1]:
                                    # print("Its", j+1,"and",j+2)
                                    j = j + 1
                            else:
                                    break
                    if i != j:
                            k = 0
                            # print(word1[j])
                            attemp.append(word1[j])
                            while k < leng - 1:
                                    data = num[k + 1]
                                    vic = num[k]
                                    # print("num", vic, "data", data)
                                    if vic == word1[j] and data == "Победа":
                                            win = win + 1
                                            # print(win)
                                    k = k + 2
                            answer = ((round(float(win / (j - i + 1)), 3)) * 100)
                            attemp.append(float(answer))
                            i = j + 1
                            # print(attemp)

                    else:
                            i = i + 1
            # print(attemp)
            return attemp

    # Функция убирающая дни, на которые выпадает ссесия а также убирает дни о которых слишком мало информации
    def blackListser(self, database):
            section = str(self.Sports_label1.currentText())
            workbook = openpyxl.load_workbook(f"Data (3).xlsx")
            sheet = workbook.active

            list = []
            list1 = []
            list2 = []
            i = 1
            j = 0
            t = 0
            c = 0
            while i < len(database):
                    # print(database[i])
                    if database[i] == 100.0 or database[i] == 0.0:
                            list.append(database[i - 1])
                            list.append(0)
                    else:
                            list.append(database[i - 1])
                            list.append(database[i])
                    i = i + 2
            for j in range(2, sheet.max_row + 1):
                    ban = sheet.cell(row=j, column=5)
                    ban = str(ban.value)
                    # print(ban)
                    ban = ban.split(' ')
                    ban = ban[0]
                    ban = ban.split('-')
                    # print(ban)
                    ban = str(ban[1] + "." + ban[2])
                    # print(ban)
                    banlist.append(ban)
            for k in range(len(banlist)):
                    n = 0
                    while n < len(database):
                            if banlist[k] == database[n]:
                                    list[n + 1] = 0
                                    # print(database[n])
                            n = n + 2

            # Лист в словарь и сделать сортироку по значениям, -
            while t < len(list):
                    list1.append(list[t])
                    list2.append(list[t + 1])
                    t = t + 2

            sortdict = dict(zip(list1, list2))

            sortdict = sorted(sortdict.items(), key=operator.itemgetter(1), reverse=True)

            # Значение + ключ в список

            return sortdict

    def algor(self, imp):
            sorter = self.sortser(imp)
            statistic = self.statisticser(sorter)
            num = statistic
            static = sorted(statistic)
            attemp = self.equalser(static, num)
            list = self.blackListser(attemp)

            return list

    def retranslateUi(self, Window):
        _translate = QtCore.QCoreApplication.translate
        self.Header.setToolTip(_translate("Window", "<html><head/><body><p align=\"center\"><br/></p></body></html>"))
        self.Header.setText(_translate("Window", "ЧЕРНЫЕ МЕДВЕДИ АНАЛИТИКА"))
        self.Label_sports.setText(_translate("Window", "Выберите секцию:"))
        self.Label_choose.setText(_translate("Window", "Выберите команду:"))
        self.Sports_label1.setItemText(0, _translate("Window", "Выберите вид спорта"))
        self.Sports_label1.setItemText(1, _translate("Window", "Волейбол"))
        self.Sports_label1.setItemText(2, _translate("Window", "Баскетбол"))
        self.Sports_label1.setItemText(3, _translate("Window", "Футбол"))
        self.Team.setItemText(0, _translate("Window", "Команда "))
        self.Team.setItemText(1, _translate("Window", "w1"))
        self.Push_done.setText(_translate("Window", "Готово"))
        self.footer.setText(_translate("Window", "СПОРТИВНЫМ БЫТЬ\n"
"       ВЕЛИКИМ БЫТЬ"))
        self.Result.setText(_translate("Window", "Лучше даты в этом месяцe:"))
        self.Full_time.setText(_translate("Window", "Процент выигрыша:"))
        self.Setting_fullname.setText(_translate("Window", "Подобрать время"))
        self.Come_back_button1.setText(_translate("Window", "Вернуться назад"))
        self.Fullname.setText(_translate("Window", "Команда:"))
        self.Add_a_student.setText(_translate("Window", "Дата:"))
        self.Sports_label2.setText(_translate("Window", "Результат:"))
        self.Choose_sports_label.setItemText(0, _translate("Window", "Победа"))
        self.Choose_sports_label.setItemText(1, _translate("Window", "Поражение"))
        self.pushButton_3.setText(_translate("Window", "Добавить студента"))
        self.Come_back_button2.setText(_translate("Window", "Вернуться назад"))


if __name__ == "__main__":
        app = QtWidgets.QApplication(sys.argv)
        MainWindow = QtWidgets.QMainWindow()
        MainWindow.setFixedSize(833, 520)
        ui = Ui_Window()
        ui.setupUi(MainWindow)
        MainWindow.show()
        app.exec()