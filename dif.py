import openpyxl
import math
import operator

statistic = []
num = []
banlist = []


workbook = openpyxl.load_workbook("Data (3).xlsx")
sheet = workbook.active

def sortser(company):
    comp = []
    for i in range(2, ws.max_row + 1) :
        date = sheet.cell(row=i, column=1)
        date = date.value
        if date == company:
             comp.append(i)
        #print(date)
        #print(comp)
    #print(comp[0])
    #print(comp[2])
    return comp
#Функция приводящая получаемые из таблицы данные в удобные для обработки значения
def statisticser(sorter):
    i = 0
    while i < len(sorter) :
        date = sheet.cell(row = sorter[i],column=2)
        date = str(date.value)
        vic = sheet.cell(row=sorter[i], column=3)
        vic = str(vic.value)

        #print(date)
        date = date.split(' ')
        date = date[0]
        date = date.split('-')
        date = str(date[1] + "." + date[2])
        #Дата в Американском фармате(позволяет использовать встроенный sort)
        statistic.append(date)
        statistic.append(vic)
        i = i + 1


    return statistic

#Функция Получения вероятности победы

def equalser(list,num) :
    attemp = []
    word1 = list
    i = 0

    leng = len(list)
    while i < (leng-1):
           j = i
           win = 0
           while(j+1 < leng ) :

               if word1[j] == "Победа" or word1[j] == "Поражение" :
                   break
               if word1[j] == word1[j+1]:
                   #print("Its", j+1,"and",j+2)
                   j = j + 1
               else:
                   break
           if i != j:
               k = 0
               #print(word1[j])
               attemp.append(word1[j])
               while k < leng - 1 :
                   data = num[k + 1]
                   vic = num[k]
                   #print("num", vic, "data", data)
                   if vic == word1[j] and data == "Победа":
                       win = win + 1
                       #print(win)
                   k = k + 2
               answer = ((round(float(win / (j - i + 1)),3)) * 100)
               attemp.append(float(answer))
               i = j + 1
               #print(attemp)

           else: i = i + 1
    #print(attemp)
    return attemp





#Функция убирающая дни, на которые выпадает ссесия а также убирает дни о которых слишком мало информации
def blackListser(database):
    list = []
    list1 = []
    list2 = []
    i = 1
    j = 0
    t = 0
    c = 0
    while i < len(database) :
        #print(database[i])
        if database[i] == 100.0 or database[i] == 0.0:
            list.append(database[i - 1])
            list.append(0)
        else:
            list.append(database[i-1])
            list.append(database[i])
        i = i + 2
    for j in range(2, sheet.max_row + 1):
        ban = sheet.cell(row=j , column=5)
        ban = str(ban.value)
        #print(ban)
        ban = ban.split(' ')
        ban = ban[0]
        ban = ban.split('-')
        #print(ban)
        ban = str(ban[1] + "." + ban[2])
        #print(ban)
        banlist.append(ban)
    for k in range(len(banlist)):
        n = 0
        while n < len(database) :
            if banlist[k] == database[n] :
                list[n+1] = 0
                #print(database[n])
            n = n + 2



    #Лист в словарь и сделать сортироку по значениям, -
    while t < len(list):
        list1.append(list[t])
        list2.append(list[t+1])
        t = t +2


    sortdict = dict(zip(list1,list2))


    sortdict = sorted(sortdict.items(), key=operator.itemgetter(1), reverse= True)

    #Значение + ключ в список


    return sortdict

def algor (imp):
    sorter = sortser(imp)
    statistic = statisticser(sorter)
    num = statistic
    static = sorted(statistic)
    attemp = equalser(static,num)
    list = blackListser(attemp)
    print(list)

